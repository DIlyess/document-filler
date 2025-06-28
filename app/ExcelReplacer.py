#!/usr/bin/env python 3.9
# -*- coding: utf-8 -*-
# @Time    : 2024/12/6 18:02 update
# @Author  : Document Filler
# @File    : ExcelReplacer.py
# @Software: PyCharm
# @Notice  : Excel file placeholder replacement while preserving formatting

import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import time


class ExcelReplace:
    """
    Excel file placeholder replacement while preserving formatting
    file: Microsoft Office Excel file, supports .xlsx and .xls files
    """

    def __init__(self, file_path):
        self.file_path = file_path
        try:
            self.workbook = load_workbook(file_path)
            self.sheets = self.workbook.sheetnames
        except Exception as e:
            raise Exception(f"Error loading Excel file {file_path}: {str(e)}")

    def replace_in_cell(self, cell, old_text, new_text):
        """
        Replace text in a cell while preserving formatting
        """
        if cell.value is None:
            return
        
        try:
            cell_value = str(cell.value)
            if old_text in cell_value:
                cell.value = cell_value.replace(old_text, new_text)
        except Exception as e:
            # Skip cells that can't be processed
            pass

    def replace_in_sheet(self, sheet_name, replace_dict):
        """
        Replace placeholders in a specific sheet
        """
        try:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    for key, value in replace_dict.items():
                        self.replace_in_cell(cell, key, value)
        except Exception as e:
            print(f"Warning: Error processing sheet {sheet_name}: {str(e)}")

    def replace_in_all_sheets(self, replace_dict):
        """
        Replace placeholders in all sheets of the workbook
        """
        for sheet_name in self.sheets:
            self.replace_in_sheet(sheet_name, replace_dict)

    def save(self, filepath=None):
        """
        Save the modified workbook
        """
        if filepath is None:
            filepath = self.file_path
        try:
            self.workbook.save(filepath)
        except Exception as e:
            raise Exception(f"Error saving Excel file {filepath}: {str(e)}")

    @staticmethod
    def excel_list(dir_path):
        """
        Get list of Excel files in directory and subdirectories
        """
        file_list = []
        for roots, dirs, files in os.walk(dir_path):
            for file in files:
                # Find Excel documents and exclude temporary files
                if (file.endswith(".xlsx") or file.endswith(".xls")) and file[0] != "~":
                    file_root = os.path.join(roots, file)
                    file_list.append(file_root)
        return file_list

    def replace_excel(self, replace_dict):
        """
        Main method to replace placeholders in Excel file
        """
        self.replace_in_all_sheets(replace_dict)
        return self.workbook

    def set_date_and_place(self):
        """
        Replace date and place placeholders in Excel file
        """
        date_placeholders = {
            "[date]": time.strftime("%d/%m/%Y"),
            "[date_du_jour]": time.strftime("%d/%m/%Y"),
            "[Fait_a]": "Arles"
        }
        self.replace_in_all_sheets(date_placeholders)


def main():
    """
    Example usage
    """
    # Example replace dictionary
    replace_dict = {
        "[NOM_ORGANISME]": "Example Organization",
        "[DATE]": "01/01/2024",
        "[ADRESSE]": "123 Example Street"
    }
    
    # Example file directory
    file_dir = "app/templates"
    
    # Process all Excel files
    for i, file in enumerate(ExcelReplace.excel_list(file_dir), start=1):
        print(f"{i}. Processing file: {file}")
        excel_replace = ExcelReplace(file)
        excel_replace.replace_excel(replace_dict)
        excel_replace.save(file)
        print(f"Excel document processing complete!")


if __name__ == "__main__":
    main() 